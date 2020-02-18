import re
import string
from datetime import date
import openpyxl as px

INVALID_ID = "Only letters and digits can comprise the {}"

# to_SQL_date(self, bday)
def to_sql_date(bday: str) -> str:
    """
    convert csv date to sql date ('YYYY-mm-dd')
    """
    # dd/mm/yyyy
    four_digit_slash_year = re.findall(r"[\d]{1,2}/[\d]{1,2}/[\d]{4}", bday)
    # dd-mm-yyyy
    four_digit_dash_year = re.findall(r"[\d]{1,2}-[\d]{1,2}-[\d]{4}", bday)
    # dd/mm/yy
    two_digit_slash_year = re.findall(r"[\d]{1,2}/[\d]{1,2}/[\d]{2}", bday)
    # dd-mm-yy
    two_digit_dash_year = re.findall(r"[\d]{1,2}-[\d]{1,2}-[\d]{2}", bday)

    if len(four_digit_slash_year) > 0:
        month, day, year = four_digit_slash_year[0].split('/')
        return f'{year}-{month}-{day}'
    elif len(four_digit_dash_year) > 0:
        month, day, year = four_digit_dash_year[0].split('-')
        return f'{year}-{month}-{day}'
    elif len(two_digit_slash_year) > 0:
        month, day, year = two_digit_slash_year[0].split('/')
        century_20_bday = int(year) + 1900
        return f'{century_20_bday}-{month}-{day}'
    elif len(two_digit_dash_year) > 0:
        month, day, year = two_digit_dash_year[0].split('-')
        century_20_bday = int(year) + 1900
        return f'{str(century_20_bday)}-{month}-{day}'
    else:
        raise Exception


def to_sql_date_alt(bday: str) -> date:
    """
    Alternate date conversion  method that uses python datetime and format modules
    :return: 
    """
    # dd/mm/yyyy
    four_digit_slash_year = re.findall(r"[\d]{1,2}/[\d]{1,2}/[\d]{4}", bday)
    # dd-mm-yyyy
    four_digit_dash_year = re.findall(r"[\d]{1,2}-[\d]{1,2}-[\d]{4}", bday)
    # dd/mm/yy
    two_digit_slash_year = re.findall(r"[\d]{1,2}/[\d]{1,2}/[\d]{2}", bday)
    # dd-mm-yy
    two_digit_dash_year = re.findall(r"[\d]{1,2}-[\d]{1,2}-[\d]{2}", bday)

    if len(four_digit_slash_year) > 0:
        month, day, year = four_digit_slash_year[0].split('/')
        four_digit_dash_year_date = date(int(year), int(month), int(day))
        return four_digit_dash_year_date
        #return f'{year}-{month}-{day}'
    elif len(four_digit_dash_year) > 0:
        month, day, year = four_digit_dash_year[0].split('-')
        four_digit_dash_year_date = date(int(year), int(month), int(day))
        return four_digit_dash_year_date
        #return f'{year}-{month}-{day}'
    elif len(two_digit_slash_year) > 0:
        month, day, year = two_digit_slash_year[0].split('/')
        century_20_bday = int(year) + 1900
        two_digit_dash_year_date = date(century_20_bday, int(month), int(day))
        return two_digit_dash_year_date
        #return f'{century_20_bday}-{month}-{day}'
    elif len(two_digit_dash_year) > 0:
        month, day, year = two_digit_dash_year[0].split('-')
        century_20_bday = int(year) + 1900
        return f'{str(century_20_bday)}-{month}-{day}'
    else:
        raise Exception


def valid_id(id) -> bool:
    """
    A unique identifier with only letters and digits
    :param id:
    :return:
    """
    ws = [ws for ws in string.whitespace if id.find(ws) is not -1]
    if len(ws) > 0:
        return False
    return True

# All of these date formats assume a four digit year
# Used by the Date class to decode different formats for dates
_format = {
    'ymd' : '{d.year}-{d.month}-{d.day}',
    'mdy' : '{d.month}/{d.day}/{d.year}',
    'dmy' : '{d.day}/{d.month}/{d.year}'
}


class Date:
    """
    The Date class is used to demonstrate another option for formatting dates
    """
    def __init__(self, year, month, day):
        """
        Construct date class
        """
        self.year = year
        self.month = month
        self.day = day

    def __format__(self, code: str):
        if code == '':
            code = 'ymd'
        elif len(re.findall(r'[\d]', code)) > 0:
            raise Exception("code must only be comprised with 'm', 'y', 'd' ")
        elif len([ws for ws in string.whitespace if code.find(ws) != -1]) > 0:
            raise Exception("code contains whitespace characters. only 'm', 'y', 'd' are allowed")

        fmt = _format[code]
        return fmt.format(d=self)

def make_date(date_str):
    if date_str.find('/') is not -1:
        y, m, d = date_str.split('/')
    elif date_str.find('-') is not -1:
        y, m, d = date_str.split('-')
    else:
        raise Exception("Illegal date format")
    return date(int(y), int(m), int(d))

def add_prez_inauguration_age(ws):
    """
    Adds the age that each president was inaugurated to a column in the presidents.xlsx file
    :param ws: presidents excel worksheet
    :return:
    """
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col).value = 'Age at Inauguration'
    for row in range(2, 47):
        birth_date = make_date(ws.cell(row=row, column=4).value)
        inaugural_date = make_date(ws.cell(row=row, column=8).value)
        raw_age_took_office = inaugural_date - birth_date
        human_age_took_office = raw_age_took_office.days / 365.25
        ws.cell(row=row, column=new_col).value = human_age_took_office


